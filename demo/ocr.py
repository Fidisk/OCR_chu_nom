from paddleocr import PaddleOCR
import os
import openpyxl

# Initialize PaddleOCR with only the recognition model
ocr = PaddleOCR(
    use_gpu=False,
    use_angle_cls=False,
    cls=False,
    det=False,
    rec=True,
    rec_model_dir="model/rec",
    lang="ch",  # Specify language (e.g., 'ch' for Chinese)
)

# Folder containing cropped single-character images
cropped_folder = "result\\output\\crop"

# Load the existing workbook (it already has the first three columns)
excel_path = "result\\output\\recognized_results.xlsx"

# Load the workbook and the active sheet
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print("Ok")

# Assuming the first row is headers, start processing from the second row
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    # Extract the Image Name (assuming the second column contains image name)
    image_name = f"{row[0].value}.jpeg"
    file_path = os.path.join(cropped_folder, image_name)
    
    if os.path.isfile(file_path):  # Ensure it's a valid file
        try:
            # Perform recognition
            result = ocr.ocr(file_path, cls=False, det=False, rec=True)
            
            # Extract the recognized text and confidence
            if result:
                recognized_text = result[0][0][0]
                confidence = result[0][0][1]
            else:
                recognized_text = ""
                confidence = 0

            print(f"{image_name} {recognized_text}")

            # Add the recognized text and confidence to the row
            ws.cell(row=row[0].row, column=4, value=recognized_text)  # Column 4 is "Recognized Text"
            ws.cell(row=row[0].row, column=5, value=confidence)      # Column 5 is "Confidence"
        
        except Exception as e:
            print(f"Error processing file {image_name}: {e}")

# Save the updated Excel file
wb.save(excel_path)
print(f"Recognition results saved to {excel_path}")
