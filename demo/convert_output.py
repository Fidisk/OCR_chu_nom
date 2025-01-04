import pandas as pd
import re
import json
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import Font
from xlsxwriter.workbook import Workbook
import Levenshtein
from itertools import combinations, permutations
from itertools import islice
# pip install python-Levenshtein for MED Levenstein 
COLUMN_THRESHOLD = 15  # Adjust this threshold based on column spacing



def load_label_data(label_file):
    with open(label_file, 'r', encoding='utf-8') as file:
        try:
            # Parse the entire JSON file
            label_data = json.load(file)
            return label_data
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON file: {e}")
            return {}
def sort_box(box_data):
    # Group boxes by columns
    box_data.sort(key=lambda box: box['points'][0][1])
    columns = []
    current_column = []
    for box in box_data:
        if not current_column:
            current_column.append(box)
        else:
            # If the x_min of the current box is close to the x_min of the last box in the column,
            # it's considered part of the same column.
            if abs(box['points'][0][1] - current_column[-1]['points'][0][1]) < COLUMN_THRESHOLD:
                current_column.append(box)
            else:
                # Sort current column by y_min (top to bottom) and add it to columns
                current_column.sort(key=lambda box: box['points'][0][0])
                columns.append(current_column)
                current_column = [box]

     # Add the last column
    if current_column:
        current_column.sort(key=lambda box: box['points'][0][0])
        columns.append(current_column)

    result = []
    for column in columns:
        for box in column:
            result.append(box)
            #print(f"{box}\n")
    return result

def find_closest_combination(chu_quoc_ngu, data, input_start,h):
    min_distance = 1000
    closest_permutation = None
    closest_bboxes = None
    start_index_of_closest = None
    data_length = len(data)
    image_name = None

    # Adjust the iteration range to start from input_start and limit to input_start + 10
    for start in range(input_start, min(input_start + h, data_length)):
        window_text = ""
        bboxes = []
        for length in range(1, 3):  # Window size: 1, 2, or 3 elements
            if start + length > data_length:  # Break if window exceeds bounds
                break

            # Add next element to the window
            current_data = data[start + length - 1]
            window_text += current_data['transcription']
            bboxes.append(current_data['points'])

            # Compute distance using MED
            try:
                distance = MED(chu_quoc_ngu.split(), window_text, equal_func=check_intersection)
            except ValueError as e:
                print(f"Caught an error: {e}")
            if distance < min_distance:
                min_distance = distance
                closest_permutation = window_text
                closest_bboxes = bboxes
                # Adjust start_index_of_closest to return the end of the current window
                start_index_of_closest = start + length  # Adjust index based on combination length
                image_name = data[start]['image_name']
                # Early exit if perfect match found
                if min_distance == 0:
                    return closest_permutation, closest_bboxes, start_index_of_closest,h
    if(len(chu_quoc_ngu.split()) <= min_distance):
        print(f"Found no OCR match for '{chu_quoc_ngu}' at index {input_start}")
        start_index_of_closest = input_start
        h += 1
    # Return the closest match and saved start index
    return closest_permutation, closest_bboxes, start_index_of_closest,image_name,h

df_quocngu = pd.read_excel('QuocNgu_SinoNom_Dic.xlsx')
df_similar = pd.read_excel('SinoNom_similar_Dic_v2.xlsx')
def read_QuocNgu(quoc_ngu_value,df):
      
    normalized_value = normalize_text(quoc_ngu_value)
    filtered_df = df[df.iloc[:, 0] == normalized_value]
    sino_nom_set = set(filtered_df.iloc[:, 1].dropna().unique())
    return sino_nom_set
def normalize_text(text):
    # Convert to lowercase, remove punctuation, and strip whitespace
    text = text.lower()
    text = re.sub(r'[^\w\s]', '', text)  
    text = text.strip()  
    return text

def read_Similar(input_char,df):
    row = df[df['Input Character'] == input_char]
    
    if not row.empty:
        similar_chars = row.iloc[0, 1:]  
        similar_set = set(similar_chars.dropna().values.flatten())
        return similar_set
    else:
        return f"No similar characters found for '{input_char}'"

def check_intersection(quoc_ngu_value, input_char):
    sino_nom_set = read_QuocNgu(quoc_ngu_value,df_quocngu)
    similar_set = read_Similar(input_char,df_similar)
    print(f"sinoNomSet of{quoc_ngu_value} is {sino_nom_set}")
    print(f"similarSet of{input_char} is {similar_set}")
    if input_char in sino_nom_set or sino_nom_set.intersection(similar_set):
        print("True")  # Check for non-empty intersection
        return True
    else:
        print("False")
        return False
def min_edit_distance_align_custom(s1, s2, equal_func,red_format):
    m, n = len(s1), len(s2)
    dp = [[0 for _ in range(n + 1)] for _ in range(m + 1)]

    # Initialize base cases
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j

    # Populate the DP table
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if equal_func(s1[i - 1], s2[j - 1]):  # Custom equality
                cost = 0
            else:
                cost = 1
            dp[i][j] = min(
                dp[i - 1][j] + 1,      # Deletion
                dp[i][j - 1] + 1,      # Insertion
                dp[i - 1][j - 1] + cost  # Substitution
            )

    # Trace back to find alignment
    i, j = m, n
    aligned_s1 = []
    aligned_s2 = []
    formatted_text = [] # For coloring the alignment
    while i > 0 or j > 0:
        if i > 0 and j > 0 and dp[i][j] == dp[i - 1][j - 1] + (0 if equal_func(s1[i - 1], s2[j - 1]) else 1):
            # Match or substitution
            if equal_func(s1[i - 1], s2[j - 1]):
                formatted_text.extend([s2[j-1]]) 
            else:
                formatted_text.extend([red_format, s2[j-1]])
            aligned_s1.append(s1[i - 1])
            aligned_s2.append(s2[j - 1])
            i -= 1
            j -= 1
        elif i > 0 and dp[i][j] == dp[i - 1][j] + 1:
            # Deletion
            formatted_text.extend([red_format,'-'])
            aligned_s1.append(s1[i - 1])
            aligned_s2.append('-')
            i -= 1
        else:
            # Insertion
            
            aligned_s1.append('-')
            aligned_s2.append(s2[j - 1])
            j -= 1

    # Reverse the alignment (as we built it backward)
    aligned_s1.reverse()
    aligned_s2.reverse()
    formatted_text.reverse()

    return dp[m][n], ' '.join(aligned_s1), ''.join(aligned_s2), formatted_text

def MED(s1, s2, equal_func):
    m, n = len(s1), len(s2)
    dp = [[0 for _ in range(n + 1)] for _ in range(m + 1)]

    # Initialize base cases
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j

    # Populate the DP table
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if equal_func(s1[i - 1], s2[j - 1]):  # Custom equality
                cost = 0
            else:
                cost = 1
            dp[i][j] = min(
                dp[i - 1][j] + 1,      # Deletion
                dp[i][j - 1] + 1,      # Insertion
                dp[i - 1][j - 1] + cost  # Substitution
            )

    

    return dp[m][n]
       
def color_OCR(data1,chu_quoc_ngu_list):
    workbook = xlsxwriter.Workbook("output.xlsx")
    worksheet = workbook.add_worksheet()

    red_format = workbook.add_format({'color': 'red'})
    blue_format = workbook.add_format({'color': 'blue'})
    green_fill = workbook.add_format({'bg_color': 'green'})

    #MED 
    data = []
    i = 0
    k = 0 
    temp = None
    start = 0
    h = 12
    h_tmp = 0
    for chu_quoc_ngu in chu_quoc_ngu_list:
        closest_permutation,closest_bboxes,end,image_name,h_tmp = find_closest_combination(chu_quoc_ngu,data1,start,h)
        h = h_tmp
        start  = end
        if(temp != image_name):
            k = 0
            temp = image_name
        composed_id = f"{image_name}_{k}"
        distance = 0
        aligned_s1 = []
        aligned_s2 = []
        formatted_text = []
        try:
            distance,aligned_s1,aligned_s2,formatted_text = min_edit_distance_align_custom(chu_quoc_ngu.split(), closest_permutation, equal_func=check_intersection,red_format=red_format)   
        except ValueError as e:
            print(f"Caught an error: {e}")
        row = {
            'ImageName': image_name,
            'ID': composed_id,
            'Box': closest_bboxes,
            'SinoNomOCR': aligned_s2,
            'QuocNgu': chu_quoc_ngu_list[i],
            'FormattedText': formatted_text,
        }
        i += 1
        k += 1
        data.append(row)

    headers = ["ImageName","ID","List Box","Sino Nom OCR","Chu Quoc Ngu"]
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header) 
    j = 0
    for box in data:
        sino_nom_ocr_text = box['SinoNomOCR']
        chu_quoc_ngu_text = box['QuocNgu'].split() if box['QuocNgu'] else []
        formatted_text = box['FormattedText']
        print(f"{sino_nom_ocr_text}")
        print(f"{chu_quoc_ngu_text}")
        list_box = ''
        for temp in box['Box']:

            flat_box = ' '.join(f"({x},{y})" for x, y in temp)
            list_box += flat_box + ';'
        if not chu_quoc_ngu_text:
            worksheet.write(j, 2,list_box , green_fill)
        else:
            worksheet.write(j, 2,list_box )
        
        

        worksheet.write_rich_string(j, 3, *formatted_text)
        worksheet.write(j, 0,box['ImageName'])
        worksheet.write(j, 1,box['ID'])
        worksheet.write(j, 4,box['QuocNgu'])
        j += 1
    workbook.close()
    print("Formatted output saved to 'output.xlsx'")


label_file = "label.json"    
label_data = load_label_data(label_file)
# Print the parsed data
index = 0
temp = []
data1 = []
for item in label_data:
    image_name = item.get('image_name')
    box_data = item.get('ocr_details')
    print(f"Image: {image_name}")
    for box in box_data:
        # print(f"  Transcription: {box['transcription']}")
        # print(f"  Points: {box['points']}")
        # print(f"  Difficult: {box['difficult']}")
        print(f"Sort box\n")
    data_item = sort_box(box_data)
    for box in data_item:
        box['image_name'] = image_name 
    temp.append(data_item)

for data_item in temp:
    for box in data_item:
        data1.append(box)
print(data1)
## MED Levenshtein for check ocr fail 
chu_quoc_ngu_list = []
with open("Bach van thi tap _Dich.txt", 'r', encoding='utf-8') as file:
    for line in file:
            #print(line.strip())
            chu_quoc_ngu_list.append(normalize_text(line.strip()))


#print("Chữ Quốc Ngữ List:", chu_quoc_ngu_list)

print(data1)
print(chu_quoc_ngu_list)

exit()
color_OCR(data1,chu_quoc_ngu_list)



    


