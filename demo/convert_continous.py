import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import fitz
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText,TextBlock,InlineFont
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
from xlsxwriter.workbook import Workbook
with open('Label.txt', 'r', encoding='utf-8') as file:
    data = file.readlines()

# Initialize lists to store parsed data
image_files = []
transcriptions = []
points = []
difficulties = []
ids = []
sinochar = []
QN = []

# Process each line in the file
for line in data:
    #print(line)
    # Split the line into the image file path and the JSON annotations
    parts = line.strip().split('\t')
    #print(parts)
    #continue
    if len(parts) != 2:
        continue  # Skip any line that doesn't have the expected format

    image_file = parts[0]
    annotations = json.loads(parts[1])

    # Iterate through each annotation for the current image
    for idx, annotation in enumerate(annotations):
        transcription = annotation.get('transcription', '')
        point = annotation.get('points', [])
        difficult = annotation.get('difficult', False)

        # Store the parsed information
        image_files.append(image_file)
        transcriptions.append(transcription)
        points.append(point)
        difficulties.append(difficult)

        # Generate a unique ID using image file name and the count of annotations
        unique_id = f"{image_file}.{idx + 1}"
        print(f"{unique_id}")
        ids.append(unique_id)

# Combine data and sort
combined = list(zip(points, ids, image_files, transcriptions, difficulties))
combined_sorted = sorted(combined, key=lambda x: (-x[0][0][0], x[0][0][1]))

# Unzip sorted data back into individual lists
points, ids, image_files, transcriptions, difficulties = zip(*combined_sorted)

eps = 20

#print(points[0][0])
#print(points[0][0][1])

points = list(points)
ids = list(ids)
image_files = list(image_files)
transcriptions = list(transcriptions)
difficulties = list(difficulties)

for i in range(len(points)):
    for j in range(i, len(points)):
        # Do something with points[i] and points[j]
        id1 = ids[i].split(".")
        fid1 = ".".join(id1[:-1])
        id2 = ids[j].split(".")
        fid2 = ".".join(id2[:-1])
        print(f"{fid1} {fid2} {fid1<fid2}")
        if fid1>fid2:
            points[i],points[j] = points[j],points[i]
            ids[i],ids[j] = ids[j],ids[i]
            image_files[i],image_files[j]=image_files[j],image_files[i]
            transcriptions[i],transcriptions[j]=transcriptions[j],transcriptions[i]
            difficulties[i],difficulties[j]=difficulties[j],difficulties[i]
        elif fid1<fid2:
            continue
        elif abs(points[i][0][0]-points[j][0][0])>eps:
            if points[i][0][0]<points[j][0][0]:
                points[i],points[j] = points[j],points[i]
                ids[i],ids[j] = ids[j],ids[i]
                image_files[i],image_files[j]=image_files[j],image_files[i]
                transcriptions[i],transcriptions[j]=transcriptions[j],transcriptions[i]
                difficulties[i],difficulties[j]=difficulties[j],difficulties[i]
        else:
            if points[i][0][1]>points[j][0][1]:
                points[i],points[j] = points[j],points[i]
                ids[i],ids[j] = ids[j],ids[i]
                image_files[i],image_files[j]=image_files[j],image_files[i]
                transcriptions[i],transcriptions[j]=transcriptions[j],transcriptions[i]
                difficulties[i],difficulties[j]=difficulties[j],difficulties[i]

for i in range(len(points)):
    print(f"{ids[i]}  {points[i]} {transcriptions[i]}")

# Function to read the file and parse the Sino characters and translations into separate arrays
def read_text_from_file(file_path):
    sinochar = []
    QN = []

    state = False

    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        
        # Iterate through the lines of the file and split into Sinochar and QN arrays
        for line in lines:
            state = not state
            # Check if the line contains Sino characters (e.g., contains Chinese characters or specific symbols)
            if state:
                sinochar.append(line.strip())  # Add Sino characters to sinochar
            else:
                QN.append(line.strip())  # Add translation or question/answer to QN

    return sinochar, QN

# Example usage
file_path = 'output.txt'  # Replace with the actual path to your text file
sinochar, QN = read_text_from_file(file_path)

# Print out the arrays for verification
print("Sino characters:")
print(sinochar)

print("\nQuestion and Answer translations:")
print(QN)


file_path = 'SinoNom_similar_Dic.xlsx'

df = pd.read_excel(file_path)

SN_similar = dict(zip(df['Input Character'], df['Top 20 Similar Characters']))

file_path = 'QuocNgu_SinoNom_Dic.xlsx'

df = pd.read_excel(file_path)

input_characters = df['QuocNgu']
similar_characters = df['SinoNom']

QN_similar = {}

for i in range(len(input_characters)):
    if input_characters[i] not in QN_similar:
        #print(input_characters[i])
        QN_similar[input_characters[i]] = []
    QN_similar[input_characters[i]].append(similar_characters[i])

for i in range(len(sinochar)):
    sinochar[i]=sinochar[i].replace(' ','')
    sinochar[i]=sinochar[i].replace('。','')

print(len(transcriptions))
print(len(sinochar))
print(len(QN))

for i in range(len(transcriptions)):
    print(f"{transcriptions[i]}")
    print(f"--------------")

print(f"++++++++++++++++++++++++++++++++")

for i in range(len(sinochar)):
    print(f"{sinochar[i]}")
    print(f"--------------")

for i in range(len(QN)):
    print(f"{QN[i]}")
    print(f"--------------")

def levene(str1, str2):
    f = [[len(str1) * len(str2) * 2 for _ in range(len(str1) + len(str2) + 1)] for _ in range(len(str2) + len(str1) + 1)]
    f[0][0]=0
    for ii in range(len(str1)):
        for ij in range(len(str2)):
            i = ii + 1
            j = ij + 1
            f[i][j]=min(f[i-1][j-1],f[i-1][j],f[i][j-1])+1
            if (str1[ii]==str2[ij]):
                f[i][j]=min(f[i][j],f[i-1][j-1])
    return f[len(str1)][len(str2)]

def leveneTrace(str1, str2):
    kq = ""
    f = [[len(str1) * len(str2) * 2 for _ in range(len(str1) + len(str2) + 1)] for _ in range(len(str2) + len(str1) + 1)]
    trace = [['' for _ in range(len(str1) + len(str2) + 1)] for _ in range(len(str2) + len(str1) + 1)]
    for ii in range(len(str1)+1):
        trace[ii][0]='_'
        f[ii][0]=ii
    for ii in range(len(str2)+1):
        trace[0][ii]=''
        f[0][ii]=ii
    f[0][0]=0
    for ii in range(len(str1)):
        for ij in range(len(str2)):
            i = ii + 1
            j = ij + 1
            #print(f"{i} {j}")
            #f[i][j]=min(f[i-1][j-1],f[i-1][j],f[i][j-1])+1
            if (str1[ii]==str2[ij]):
                if f[i-1][j-1] == min(f[i-1][j-1], f[i][j-1] + 1, f[i-1][j] + 1):
                    f[i][j] = f[i-1][j-1]
                    trace[i][j] = str2[ij]
                elif f[i][j-1] + 1 == min(f[i-1][j-1], f[i][j-1] + 1, f[i-1][j] + 1):
                    f[i][j] = f[i][j-1] + 1
                    trace[i][j] = ''
                else:
                    f[i][j] = f[i-1][j] + 1
                    trace[i][j] = '_'
            else:
                if f[i-1][j-1] + 1 == min(f[i-1][j-1] + 1, f[i][j-1] + 1, f[i-1][j] + 1):
                    f[i][j] = f[i-1][j-1] + 1
                    trace[i][j] = str2[ij]
                elif f[i][j-1] + 1 == min(f[i-1][j-1] + 1, f[i][j-1] + 1, f[i-1][j] + 1):
                    f[i][j] = f[i][j-1] + 1
                    trace[i][j] = ''
                else:
                    f[i][j] = f[i-1][j] + 1
                    trace[i][j] = '_'
    tracex = len(str1)
    tracey = len(str2)
    #print(f"{str1} {str2}")
    while tracex != 0 or tracey != 0:
        kq+=trace[tracex][tracey]
        if tracey == 0:
            tracex -= 1
        elif tracex == 0:
            tracey -= 1
        else:
            if trace[tracex][tracey]=='_':
                tracex -= 1
            elif trace[tracex][tracey] == '':
                tracey -= 1
            else:
                tracex -= 1
                tracey -= 1

    #for ii in range(len(str1)+1):
    #    print(f"{f[ii][0:len(str2)+1]}")
    kq=''.join(list(reversed(kq)))
    return kq

def similar(str1, str2):
    return len(str2)-levene(str1,str2)>=len(str1)/3

cost = 0

sinochar2 = sinochar
sinochar = []
QN2 = QN
QN = []

for i in range(len(transcriptions)):
    transcriptions[i]=transcriptions[i].replace(' ','')
    transcriptions[i]=transcriptions[i].replace("。","")

for i in range(len(QN2)):
    QN2[i] = QN2[i].replace('.','')
    QN2[i] = QN2[i].replace('!','')
    QN2[i] = QN2[i].replace('?','')
    QN2[i] = QN2[i].replace(',','')
    QN2[i] = QN2[i].replace('-','')
    QN2[i] = QN2[i].replace("  ",' ')
    QN2[i] = QN2[i].replace("17","")

for i in range(len(QN2)):
    QN2[i] = QN2[i].replace('.','')
    QN2[i] = QN2[i].replace('!','')
    QN2[i] = QN2[i].replace('?','')
    QN2[i] = QN2[i].replace(',','')
    QN2[i] = QN2[i].replace('-','')
    QN2[i] = QN2[i].replace("  ","")
    QN2[i] = QN2[i].replace("17","")

msinochar = ''
mQN = ''
mtrans = ''

for i in range(len(QN2)):
    if (mQN != ''):
        mQN = mQN + ' ' + QN2[i]
    else:
        mQN = QN2[i]

lQN = mQN.split(' ')

for i in range(len(sinochar2)):
    msinochar+=sinochar2[i]

for i in range(len(transcriptions)):
    mtrans+=transcriptions[i]

#print(f"+++++++++++++++++++++++++++++++++++++++++++++")
#print(leveneTrace("abbbce","acbce"))
#print(leveneTrace(mtrans,msinochar))
#print(f"+++++++++++++++++++++++++++++++++++++++++++++")

msinochar=leveneTrace(mtrans,msinochar)

with open("aa.txt", "w", encoding='utf-8') as f:
    print(msinochar, file=f)

with open("bb.txt", "w", encoding='utf-8') as f:
    print(mtrans, file=f)


p = 0

for i in range(len(transcriptions)):
    QN.append('')
    sinochar.append('')
    for j in range(len(transcriptions[i])):
        #print(f"{p} | {len(lQN)}")
        if QN[i] == '':
            if p < len(lQN):
                QN[i] += lQN[p]
            else:
                QN[i] += "không"
        else:
            if p < len(lQN):
                QN[i] += ' ' + lQN[p]
            else:
                QN[i] += " không"
        if p in range(len(msinochar)):
            sinochar[i] += msinochar[p]
        p += 1

for i in range(len(transcriptions)):
    print(f"{transcriptions[i]} | {sinochar[i]} |{QN[i]}|")
    print(f"--------------")

print(f"++++++++++++++++++++++++++++++++")

#for i in range(len(transcriptions)):
    #print(f"{transcriptions[i]} {sinochar[i]}")

df = pd.DataFrame({
    'ID': ids,
    'Image Box': points,
    'SinoNom OCR': transcriptions,
    'SinoNom Char': sinochar,
    'QN': QN,
})

output_file = 'output_labels.xlsx'
df.to_excel(output_file, index=False)

workbook = load_workbook(output_file)
sheet = workbook.active

nom_na_tong_font = Font(name="Nom Na Tong")

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.font = nom_na_tong_font

workbook.save(output_file)

#Coloring


for i in range(len(transcriptions)):
    SNls = list(str(transcriptions[i]))
    QNls = str(QN[i].lower().replace(",", "").replace(".", "").replace(":", "")).split()
    charSinom = list(str(sinochar[i]))

    f = [[len(SNls) * len(QNls) * 2 for _ in range(len(SNls) + 5)] for _ in range(len(QNls) + 5)]
    trace = [['' for _ in range(len(SNls) + 5)] for _ in range(len(QNls) + 5)]
    f[0][0]=0
    #print(SNls)
    #print(QNls)
    for j in range(len(SNls)):
        for l in range(len(QNls)):
            #print(QNls[l])
            if SNls[j] not in SN_similar:
                SN_similar_ls = list([])
            else:
                SN_similar_ls = list(SN_similar[SNls[j]])
            SN_similar_ls.insert(0,SNls[j])
            if QNls[l].lower().replace(",", "").replace(".", "").replace(":", "") not in QN_similar:
                QN_similar_ls = []
            else:
                QN_similar_ls = QN_similar[QNls[l].lower().replace(",", "").replace(".", "").replace(":", "")]
            intersection = list(set(SN_similar_ls) & set(QN_similar_ls))
            #if (i == 0):
                #print("--------------------")
                    #print(SN_similar_ls)
                    #print(QN_similar_ls)
                #print(i,' ',j,' ',l)
                #print(intersection)
                #print("--------------------")
            if len(intersection) > 0:
                #print(j,' ',l)
                if f[j][l] == min(f[j][l], f[j + 1][l] + 1, f[j][l + 1] + 1):
                    f[j + 1][l + 1] = f[j][l]
                    trace[j + 1][l + 1] = intersection[0]
                elif f[j + 1][l] + 1 == min(f[j][l], f[j + 1][l] + 1, f[j][l + 1] + 1):
                    f[j + 1][l + 1] = f[j + 1][l] + 1
                    trace[j + 1][l + 1] = '-'
                else:
                    f[j + 1][l + 1] = f[j + 1][l] + 1
                    trace[j + 1][l + 1] = SNls[j]
            else:
                if f[j][l]+2 == min(f[j][l]+2, f[j + 1][l] + 1, f[j][l + 1] + 1):
                    f[j + 1][l + 1] = f[j][l]+2
                    trace[j + 1][l + 1] = SNls[j]
                elif f[j + 1][l] + 1 == min(f[j + 1][l] + 1, f[j][l + 1] + 1):
                    f[j + 1][l + 1] = f[j + 1][l] + 1
                    trace[j + 1][l + 1] = '-'
                else:
                    f[j + 1][l + 1] = f[j + 1][l] + 1
                    trace[j + 1][l + 1] = SNls[j]
    
    #print(f)
    tracex = len(SNls)
    tracey = len(QNls)
    transcriptions[i]=""
    while tracex != 0 or tracey != 0:
        transcriptions[i]+=trace[tracex][tracey]
        if tracey == 0:
            tracex -= 1
        elif tracex == 0:
            tracey -= 1
        else:
            if f[tracex][tracey] == f[tracex - 1][tracey]:
                tracex -= 1
            elif f[tracex][tracey] == f[tracex][tracey - 1]:
                tracey -= 1
            else:
                tracex -= 1
                tracey -= 1
    transcriptions[i]=''.join(list(reversed(transcriptions[i])))



df = pd.DataFrame({
    'ID': ids,
    'Image Box': points,
    'SinoNom OCR': transcriptions,
    'SinoNom Char': sinochar,
    'QN': QN,
})

output_file = 'output_labels.xlsx'
df.to_excel(output_file, index=False)

wb = xw.Book(output_file)
ws = wb.sheets[0]

for i in range(len(transcriptions)):
    cell = ws.range('C' + str(i + 2))
    SNls = transcriptions[i]
    QNls = str(QN[i].lower().replace(",", "").replace(".", "").replace(":", "")).split()
    charSinom = list(str(sinochar[i]))

    for j in range(len(SNls)):
        if j >= len(QNls):
            break

        cleaned_QN = QNls[j].lower().replace(",", "").replace(".", "").replace(":", "")
        if cleaned_QN not in QN_similar:
            QN_similar_ls = []
        else:
            QN_similar_ls = QN_similar[cleaned_QN]

        if SNls[j] in QN_similar_ls:
            if SNls[j] == charSinom[j]:
                cell.characters[j:j+1].font.color = (0, 0, 0)  # No change, keep black
            else:
                #print("Difference found: ", i, ' ', j, ' ', SNls[j], ' ', charSinom[j])
                cell.characters[j:j+1].font.color = (0, 0, 255)  # Change to blue
        else:
            if SNls[j] not in SN_similar:
                SN_similar_ls = []
            else:
                SN_similar_ls = list(SN_similar[SNls[j]])

            SN_similar_ls.insert(0,SNls[j])
            intersection = list(set(SN_similar_ls) & set(QN_similar_ls))
            if len(intersection) > 0:
                #print("Not belong to set: ", i, ' ', j, ' ', SNls[j])
                cell.characters[j:j+1].font.color = (0, 0, 255)  # Change to blue
            else:
                #print("Difference found, not belong to set: ", i, ' ', j, ' ', SNls[j], ' ', charSinom[j])
                cell.characters[j:j+1].font.color = (255, 0, 0)  # Change to red

    # Ensure the font is set for the entire cell after formatting
    cell.font.name = 'Nom Na Tong'

xw.apps.active.api.Calculate()
xw.apps.active.api.ScreenUpdating = True 

wb.save()
wb.close()

workbook = Workbook('output.xlsx')
worksheet = workbook.add_worksheet()

red = workbook.add_format({
    'font_name': 'Nom Na Tong',  # Font name
    'font_size': 18,       # Font size
    'bold': False,          # Bold font
    'italic': False,       # Italic font (optional)
    'color': 'red'
})

blue = workbook.add_format({
    'font_name': 'Nom Na Tong',  # Font name
    'font_size': 18,       # Font size
    'bold': False,          # Bold font
    'italic': False,       # Italic font (optional)
    'color': 'blue'
})

font_format = workbook.add_format({
    'font_name': 'Nom Na Tong',  # Font name
    'font_size': 18,       # Font size
    'bold': False,          # Bold font
    'italic': False,       # Italic font (optional)
    'font_color': 'black'   # Font color
})

worksheet.write(0,0,'ID',font_format)
worksheet.write(0,1,'Image Box',font_format)
worksheet.write(0,2,'SinoNom OCR',font_format)
worksheet.write(0,3,'SinoNom Char',font_format)
worksheet.write(0,4,'Chữ Quốc Ngữ',font_format)

for i in range(len(transcriptions)):
    worksheet.write(i+1,0,id[i],font_format)
    worksheet.write(i+1,1,str(points[i]),font_format)
    worksheet.write(i+1,3,sinochar[i],font_format)
    worksheet.write(i+1,4,QN[i],font_format)

for i in range(len(transcriptions)):
    format_pairs = []
    SNls = transcriptions[i]
    QNls = str(QN[i].lower().replace(",", "").replace(".", "").replace(":", "")).split()
    charSinom = list(str(sinochar[i]))

    for j in range(len(SNls)):
        if j >= len(QNls):
            break

        cleaned_QN = QNls[j].lower().replace(",", "").replace(".", "").replace(":", "")
        if cleaned_QN not in QN_similar:
            QN_similar_ls = []
        else:
            QN_similar_ls = QN_similar[cleaned_QN]

        if SNls[j] == charSinom[j]:
            format_pairs.extend((font_format,SNls[j]))
        elif SNls[j] in QN_similar_ls:
            if SNls[j] == charSinom[j]:
                format_pairs.extend((font_format,SNls[j]))
            else:
                #print("Difference found: ", i, ' ', j, ' ', SNls[j], ' ', charSinom[j])
                format_pairs.extend((blue, SNls[j]))
        else:
            if SNls[j] not in SN_similar:
                SN_similar_ls = []
            else:
                SN_similar_ls = list(SN_similar[SNls[j]])

            SN_similar_ls.insert(0, SNls[j])
            intersection = list(set(SN_similar_ls) & set(QN_similar_ls))
            if len(intersection) > 0:
                #print("Not belong to set: ", i, ' ', j, ' ', SNls[j])
                format_pairs.extend((blue, SNls[j]))
            else:
                #print("Difference found, not belong to set: ", i, ' ', j, ' ', SNls[j], ' ', charSinom[j])
                format_pairs.extend((red, SNls[j]))
    worksheet.write_rich_string(i+1, 2, *format_pairs)

workbook.close()

